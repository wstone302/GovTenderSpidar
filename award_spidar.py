from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from PIL import Image
import imagehash
import io, hashlib, time
import pandas as pd
import cv2
import numpy as np
import re
from openpyxl import load_workbook

def img_obj_to_cv(img_obj):
    return cv2.cvtColor(np.array(img_obj), cv2.COLOR_RGB2BGR)

def template_match(a_img, b_img, threshold=0.9):
    img = img_obj_to_cv(a_img)
    tpl = img_obj_to_cv(b_img)
    res = cv2.matchTemplate(img, tpl, cv2.TM_CCOEFF_NORMED)
    _, max_val, _, _ = cv2.minMaxLoc(res)
    return max_val >= threshold

def split_a_image(img_bytes):
    img = Image.open(io.BytesIO(img_bytes))
    w, h = img.size
    pad = 4  # 寬度83px，但內容71px，適當內縮
    # 左牌
    crop1 = img.crop((pad, 0, 83-pad, h))
    # 右牌
    crop2 = img.crop((83+pad, 0, w-pad, h))
    return [crop1, crop2]

def solve_card_captcha(driver, timeout=15):
    import time, io
    from PIL import Image
    import numpy as np
    import cv2
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait

    def img_obj_to_cv(img_obj):
        return cv2.cvtColor(np.array(img_obj), cv2.COLOR_RGB2BGR)

    def template_match(a_img, b_img, threshold=0.90):
        img = img_obj_to_cv(a_img)
        tpl = img_obj_to_cv(b_img)
        res = cv2.matchTemplate(img, tpl, cv2.TM_CCOEFF_NORMED)
        _, max_val, _, _ = cv2.minMaxLoc(res)
        return max_val

    wait = WebDriverWait(driver, timeout)
    all_tables = driver.find_elements(By.XPATH, "//img[contains(@src,'/tps/validate/init?poker=')]/ancestor::table[1]")
    if len(all_tables) < 2:
        print("❗ 無法定位 A/B區，請檢查HTML結構")
        return False

    a_table, b_table = all_tables[0], all_tables[1]
    a_img_ele = a_table.find_element(By.XPATH, ".//img[contains(@src,'/tps/validate/init?poker=answer')]")
    a_img_bytes = a_img_ele.screenshot_as_png
    a_imgs = split_a_image(a_img_bytes)
    b_labels = b_table.find_elements(By.XPATH, ".//label")
    b_imgs = []
    for label in b_labels:
        b_img = label.find_element(By.TAG_NAME, "img")
        while not b_img.is_displayed():
            time.sleep(0.05)
        b_img_obj = Image.open(io.BytesIO(b_img.screenshot_as_png))
        b_imgs.append(b_img_obj)

    # 產生所有A-B模板比對分數表
    score_table = np.zeros((len(a_imgs), len(b_imgs)))
    for i, a_img in enumerate(a_imgs):
        for j, b_img in enumerate(b_imgs):
            score_table[i, j] = template_match(a_img, b_img, threshold=0.9)

    # 對於每個A區，找到分數最高的B區
    chosen_j = set()
    chosen_labels = []
    for i in range(len(a_imgs)):
        # 每一A區找到匹配分數最高且沒被選過的B區
        match_j = np.argmax(score_table[i])
        if score_table[i, match_j] > 0.85 and match_j not in chosen_j:
            chosen_j.add(match_j)
            chosen_labels.append(b_labels[match_j])
            print(f"✔️ A區第{i+1}張 → B區第{match_j+1}張，分數 {score_table[i, match_j]:.2f}")

    # 點選所有正確的B區label
    for label in chosen_labels:
        try:
            label.click()
        except:
            checkbox_id = label.get_attribute("for")
            checkbox = b_table.find_element(By.ID, checkbox_id)
            checkbox.click()

    if len(chosen_labels) != len(a_imgs):
        print("❗ 沒有對到A區兩張所有對應B區")
        return False

    time.sleep(0.3)
    submit_btn = driver.find_element(By.ID, "b_submit")
    driver.execute_script("arguments[0].scrollIntoView();", submit_btn)
    submit_btn.click()
    print("✔️ 已點擊確認送出")
    time.sleep(1)
    return True

def extract_amount(text):
    """從任意金額描述只取純數字（允許有逗號），例如 '15,768,000元' -> '15768000'"""
    match = re.search(r'[\d,]+', text.replace(' ', ''))
    if match:
        return match.group(0).replace(',', '')
    return None

def crawl_detail_page(driver):
    '''擷取標案細節頁五大欄位，回傳 dict'''
    wait = WebDriverWait(driver, 10)
    target_fields = {
        "品項名稱": None,
        "得標廠商": None,
        "得標廠商原始投標金額": None,
        "決標金額": None,
        "底價金額": None
    }
    rows = driver.find_elements(By.XPATH, "//tr")
    for row in rows:
        try:
            tds = row.find_elements(By.TAG_NAME, "td")
            if len(tds) >= 2:
                label = tds[0].text.strip()
                value = tds[1].text.strip()
                for key in target_fields.keys():
                    if label == key:
                        # 金額類只存阿拉伯數字
                        if "金額" in key:
                            target_fields[key] = extract_amount(value)
                        else:
                            target_fields[key] = value
        except:
            continue
    print("📋 擷取結果：", target_fields)
    return target_fields

def main():
    # Excel 路徑與起始欄
    excel_path = "114年採購網標案投標評估彙整表 (20250722).xlsx"
    df = pd.read_excel(excel_path)
    wb = load_workbook(excel_path)
    ws = wb["採購網標案彙整表-1"]

    # 欄位對應
    start_col = 19  # S 欄
    detail_keys = ["品項名稱", "得標廠商", "得標廠商原始投標金額", "決標金額", "底價金額"]

    # 啟動瀏覽器
    driver = webdriver.Chrome()
    wait = WebDriverWait(driver, 15)
    driver.get("https://web.pcc.gov.tw/prkms/tender/common/bulletion/indexBulletion")
    time.sleep(1)

    # 只初始化一次（首筆時切換招標/決標）
    checkbox = wait.until(EC.element_to_be_clickable((By.ID, "scop1")))
    if checkbox.is_selected():
        checkbox.click()
        print("✔️ 已自動勾選『招標』")
    checkbox = wait.until(EC.element_to_be_clickable((By.ID, "scop2")))
    if not checkbox.is_selected():
        checkbox.click()
        print("✔️ 已自動勾選『決標』")

    for idx, row in df.iterrows():
        kw = str(row['案名']).strip()
        if kw.endswith("(開口合約)"):
            kw = kw[:-6].strip()
        elif kw.endswith("開口合約"):
            kw = kw[:-4].strip()
        print(f"\n🔍 查詢案名：{kw}")
        try:
            # 1. 一律直接回首頁，不用 back()
            driver.get("https://web.pcc.gov.tw/prkms/tender/common/bulletion/indexBulletion")
            time.sleep(1)
            # 2. 勾選只要「決標」沒被勾就點，招標有勾就取消
            checkbox = wait.until(EC.element_to_be_clickable((By.ID, "scop1")))
            if checkbox.is_selected():
                checkbox.click()
            checkbox = wait.until(EC.element_to_be_clickable((By.ID, "scop2")))
            if not checkbox.is_selected():
                checkbox.click()

            # 3. 搜尋輸入
            input_box = wait.until(EC.presence_of_element_located((By.NAME, "querySentence")))
            input_box.clear()
            input_box.send_keys(kw)
            driver.find_element(By.XPATH, "//a[@title='查詢']").click()
            time.sleep(2)

            # 只抓第一筆
            wait.until(EC.visibility_of_element_located((By.XPATH, "//table[@id='bulletion']/tbody/tr")))
            tr = driver.find_element(By.XPATH, "//table[@id='bulletion']/tbody/tr")
            cols = tr.find_elements(By.TAG_NAME, "td")
            if len(cols) >= 4:
                link = cols[3].find_element(By.TAG_NAME, "a")
                link.click()
                time.sleep(1)
                # 撲克牌驗證
                for _ in range(3):
                    if "撲克牌" in driver.page_source or "請於B區挑選與A區相同之撲克牌後送出" in driver.page_source:
                        ok = solve_card_captcha(driver)
                        if ok: break
                        driver.refresh(); time.sleep(1)
                    else:
                        break
                time.sleep(2)
                details = crawl_detail_page(driver)
                for k, key in enumerate(detail_keys):
                    ws.cell(row=idx+2, column=start_col+k, value=details.get(key, None))
            else:
                print(f"❗ 查無資料：{kw}")
        except Exception as e:
            print(f"❗ 查詢失敗：{e}")
            continue

    driver.quit()
    wb.save(excel_path)
    print("✅ 全部回填完畢！")

if __name__ == "__main__":
    main()
