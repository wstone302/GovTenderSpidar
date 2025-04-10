from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

#  keywords list
keywords1 = [
    "測量", "測繪", "檢測", "地形", "地籍圖", "地籍", "套疊", "套繪", "圖資", "水深", "淤積", "海床",
    "空拍", "航拍", "無人機", "飛行載具", "UAV", "UAS", "圖根點", "監測", "斷面", "掃描", "點雲",
    "LIDAR", "BIM", "光達", "沉陷", "水庫", "雷射", "疏濬", "疏浚", "地形收方", "排水", "電塔", "湖",
    "多音束", "單音束", "SBES", "離岸風電", "透地雷達", "聲納", "數值地形", "建模"
]

keywords2 = [
    "經濟部水利署北區水資源分署", "大觀發電廠", "萬大發電廠", "第三河川分署"
]

driver = webdriver.Chrome()
wait = WebDriverWait(driver, 15)
all_data = []

def parse_table_rows(rows, keyword, keyword_field):
    for row in rows:
        cols = row.find_elements(By.TAG_NAME, "td")
        if len(cols) >= 9:
            try:
                case_no = cols[2].text.strip().split()[0]
                title = cols[2].find_element(By.TAG_NAME, "a").text.strip()
                link = cols[2].find_element(By.TAG_NAME, "a").get_attribute("href")
            except:
                case_no, title, link = "", "", ""

            try:
                budget = cols[8].find_element(By.TAG_NAME, "span").text.strip()
            except:
                budget = cols[8].get_attribute("textContent").strip()

            all_data.append({
                "來源關鍵字": keyword,
                "關鍵字欄位": keyword_field,
                "案號": case_no,
                "標案名稱": title,
                "機關名稱": cols[1].text.strip(),
                "決標方式": cols[4].text.strip(),
                "採購性質": cols[5].get_attribute("textContent").strip(),
                "公告日期": cols[6].get_attribute("textContent").strip(),
                "截止日期": cols[7].get_attribute("textContent").strip(),
                "預算金額": budget,
                "詳細頁連結": link
            })

def parse_all_pages(keyword, keyword_field):
    while True:
        try:
            wait.until(EC.presence_of_element_located((By.XPATH, "//table[@id='tpam']/tbody/tr")))
            rows = driver.find_elements(By.XPATH, "//table[@id='tpam']/tbody/tr")
            parse_table_rows(rows, keyword, keyword_field)
            print(f"擷取 {len(rows)} 筆")
        except:
            print("⚠ 查無資料")
            break

        try:
            next_page = wait.until(EC.presence_of_element_located((By.LINK_TEXT, "下一頁")))
            driver.execute_script("arguments[0].scrollIntoView(true);", next_page)
            time.sleep(0.5)
            next_page.click()
            time.sleep(1)
        except:
            break

# 查詢標案名稱
for keyword in keywords1:
    print(f"\n 查詢標案名稱關鍵字：{keyword}")
    driver.get("https://web.pcc.gov.tw/prkms/tender/common/basic/indexTenderBasic")
    wait.until(EC.presence_of_element_located((By.NAME, "tenderName"))).clear()
    driver.find_element(By.NAME, "tenderName").send_keys(keyword)
    Select(wait.until(EC.presence_of_element_located((By.ID, "tenderTypeSelect")))).select_by_visible_text("招標公告")

    # 勾選「等標期限內」
    label = wait.until(EC.element_to_be_clickable((By.XPATH, "//label[@for='level_22']")))
    label.click()

    wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@title='查詢']"))).click()
    parse_all_pages(keyword, "標案名稱")
    time.sleep(1)

# 查詢機關名稱
for keyword in keywords2:
    print(f"\n 查詢機關名稱關鍵字：{keyword}")
    driver.get("https://web.pcc.gov.tw/prkms/tender/common/basic/indexTenderBasic")
    wait.until(EC.presence_of_element_located((By.NAME, "orgName"))).clear()
    driver.find_element(By.NAME, "orgName").send_keys(keyword)
    Select(wait.until(EC.presence_of_element_located((By.ID, "tenderTypeSelect")))).select_by_visible_text("招標公告")

    # 勾選「等標期限內」
    label = wait.until(EC.element_to_be_clickable((By.XPATH, "//label[@for='level_22']")))
    label.click()

    wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@title='查詢']"))).click()
    parse_all_pages(keyword, "機關名稱")
    time.sleep(1)

driver.quit()

# 資料整理與寫入
df = pd.DataFrame(all_data)
excel_path = "C:/Users/user/Desktop/spidar/政府採購標案彙整.xlsx"

if os.path.exists(excel_path):
    old_df = pd.read_excel(excel_path)
    combined_df = pd.concat([old_df, df], ignore_index=True)
    combined_df.drop_duplicates(subset=["案號", "標案名稱"], keep='last', inplace=True)
else:
    combined_df = df

# 固定欄位順序
column_order = [
    "來源關鍵字", "關鍵字欄位", "案號", "標案名稱", "機關名稱",
    "決標方式", "採購性質", "公告日期", "截止日期",
    "預算金額", "詳細頁連結"
]
combined_df = combined_df[[col for col in column_order if col in combined_df.columns]]
combined_df.to_excel(excel_path, index=False)

# 用紅色標記「更正公告」
wb = load_workbook(excel_path)
ws = wb.active
red_font = Font(color="FF0000")

# 找出「標案名稱」欄位 index
title_col_idx = None
for col_idx, cell in enumerate(ws[1], start=1):
    if cell.value == "標案名稱":
        title_col_idx = col_idx
        break

# 若標案名稱包含「更正公告」則整列標紅
if title_col_idx:
    for row in ws.iter_rows(min_row=2):
        cell = row[title_col_idx - 1]
        if "更正公告" in str(cell.value):
            for c in row:
                c.font = red_font

wb.save(excel_path)
print("✅ 任務完成，所有資料寫入並完成紅色標記。")
